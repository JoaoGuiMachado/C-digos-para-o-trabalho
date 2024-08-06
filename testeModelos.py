import FreeSimpleGUI as sg
import openpyxl
from openpyxl import load_workbook

sg.theme('Black')

# Função para abrir um arquivo Excel
def abrir_arquivo_excel(caminho_arquivo):
    try:
        if not caminho_arquivo.lower().endswith('.xlsx'):
            sg.popup_error("O arquivo deve ter a extensão .xlsx")
            return None
        return load_workbook(filename=caminho_arquivo)
    except Exception as e:
        sg.popup_error(f"Erro ao abrir o arquivo: {e}")
        return None

# Função para salvar um arquivo Excel
def salvar_arquivo_excel(workbook, caminho_arquivo):
    try:
        if not caminho_arquivo.lower().endswith('.xlsx'):
            sg.popup_error("O arquivo deve ter a extensão .xlsx")
            return
        workbook.save(caminho_arquivo)
    except Exception as e:
        sg.popup_error(f"Erro ao salvar o arquivo: {e}")

# Função para coletar dados e inserir na planilha de destino
def processar_e_salvar_dados(sheet_origem, sheet_destino):
    dados = []
    
    for row in range(1, 5001):  # Ajuste o limite conforme necessário
        for column in range(1, 31):
            celula = sheet_origem.cell(row, column)
            if celula.value:
                valor_string = str(celula.value).upper()

                if valor_string in ["RECURSOS HUMANOS", "RECURSO HUMANO", "SERVIÇO DE TERCEIROS", "SERVIÇOS DE TERCEIROS",
                                    "MATERIAL PERMANENTE", "MATERIAL E EQUIPAMENTO", "MATERIAL DE CONSUMO",
                                    "VIAGENS E DIÁRIAS", "OUTROS"]:
                    
                    dados.append((valor_string, row, column))
    
    # Ordenar dados conforme desejado (exemplo: por linha e coluna)
    dados.sort(key=lambda x: (x[1], x[2]))  # Ordena por linha e depois por coluna

    # Inserir dados na planilha de destino
    row_index_destino = 1  # Linha inicial na planilha de destino
    for valor_string, row, column in dados:
        sheet_destino.cell(row=row_index_destino, column=1, value=valor_string)  # Ajuste a coluna conforme necessário
        sheet_destino.cell(row=row_index_destino, column=2, value=row)  # Adiciona a linha
        sheet_destino.cell(row=row_index_destino, column=3, value=column)  # Adiciona a coluna
        row_index_destino += 1

def main():
    janela_principal = [  
        [sg.Text("Qual seria o nome do arquivo origem? (coloque o caminho)")],
        [sg.InputText(key="arquivo_Origem")],
        [sg.Text("Qual seria o nome do arquivo destino? (coloque o caminho)")],
        [sg.InputText(key="arquivo_Destino")],
        [sg.Button("P&D")], [sg.Button("PE")], [sg.Button("Cancelar")] 
    ]

    window_principal = sg.Window("Dados", janela_principal)

    while True:
        event_principal, values_principal = window_principal.read()
        window_principal.close()

        if event_principal == sg.WIN_CLOSED or event_principal == 'Cancelar':
            break

        arquivoDeOrigem = values_principal["arquivo_Origem"].strip().replace('"', '')
        arquivoDeDestino = values_principal["arquivo_Destino"].strip().replace('"', '')

        if event_principal == 'PE':
            print("Chegou no PE")

            workbookOrigem = abrir_arquivo_excel(arquivoDeOrigem)
            if not workbookOrigem:
                continue

            janela_escolha = [
                [sg.Text("Qual seria a aba?")],
                [sg.Text(workbookOrigem.sheetnames)],
                [sg.InputText(key='escolha_Aba')],
                [sg.Button('Usar esta aba!')], [sg.Button('Cancelar')]
            ]

            window_escolha = sg.Window("Escolha", janela_escolha)

            while True:
                event_escolha, values_escolha = window_escolha.read()
                window_escolha.close()

                if event_escolha == sg.WIN_CLOSED or event_escolha == 'Cancelar':
                    break

                if event_escolha == 'Usar esta aba!':
                    abaEscolhida = values_escolha['escolha_Aba']
                    if abaEscolhida not in workbookOrigem.sheetnames:
                        sg.popup_error("A aba escolhida não existe.")
                        continue

                    sheet_origem = workbookOrigem[abaEscolhida]

                    # Abrir ou criar o arquivo de destino
                    workbookDestino = abrir_arquivo_excel(arquivoDeDestino)
                    if not workbookDestino:
                        workbookDestino = openpyxl.Workbook()
                    
                    sheet_destino = workbookDestino.active
                    sheet_destino.title = "Dados Processados"

                    processar_e_salvar_dados(sheet_origem, sheet_destino)
                    salvar_arquivo_excel(workbookDestino, arquivoDeDestino)
                    sg.popup("Dados processados e salvos com sucesso!")
                    break

if __name__ == "__main__":
    main()
