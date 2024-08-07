import FreeSimpleGUI as sg
from openpyxl import load_workbook, Workbook
import joblib  # Para carregar o modelo treinado

# Carregue o modelo treinado
model = joblib.load('modelo_classificacao.pkl')  # Exemplo com modelo de scikit-learn
vectorizer = joblib.load('vetorizador.pkl')  # Carregue o vetorizer se necessário

# Função para processar os arquivos com IA
def process_files(source_file, dest_file, filter_text):
    # Leia o arquivo de origem
    workbook = load_workbook(filename=source_file)
    sheet = workbook.active

    # Crie um novo workbook para o arquivo de destino
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # Iterar sobre as linhas e aplicar filtro com o modelo
    for row in sheet.iter_rows(values_only=True):
        # Prepara a linha para a predição
        row_text = " ".join(str(cell) for cell in row)
        row_vector = vectorizer.transform([row_text])
        prediction = model.predict(row_vector)

        if prediction == 1:  # Suponha que '1' seja a classe relevante
            new_sheet.append(row)

    # Salve o arquivo de destino
    new_workbook.save(dest_file)

# Layout da interface gráfica
layout = [
    [sg.Text("Selecionar Arquivo de Origem"), sg.Input(), sg.FileBrowse(key="source_file")],
    [sg.Text("Selecionar Arquivo de Destino"), sg.Input(), sg.FileSaveAs(key="dest_file")],
    [sg.Text("Digite o filtro:"), sg.Input(key="filter_text")],
    [sg.Button("Aplicar Filtros e Processar"), sg.Exit()]
]

# Criação da janela
window = sg.Window("Processador de Planilhas com IA", layout)

# Loop de eventos
while True:
    event, values = window.read()

    if event in (sg.WIN_CLOSED, 'Exit'):
        break

    if event == 'Aplicar Filtros e Processar':
        source_file = values["source_file"]
        dest_file = values["dest_file"]
        filter_text = values["filter_text"]

        if not source_file or not dest_file:
            sg.popup("Por favor, selecione ambos os arquivos.")
            continue

        process_files(source_file, dest_file, filter_text)
        sg.popup(f"Dados filtrados foram salvos em: {dest_file}")

window.close()
