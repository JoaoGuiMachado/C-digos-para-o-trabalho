import FreeSimpleGUI as sg
import openpyxl
from openpyxl import load_workbook

sg.theme('Black')

# Função para abrir um arquivo Excel
def abrir_arquivo_excel(caminho_arquivo):
    try:
        if not caminho_arquivo.lower().endswith('.xlsx'):
            sg.popup_error("O arquivo deve ter a extensão .xlsx")
            return None
        return load_workbook(filename=caminho_arquivo)
    except Exception as e:
        sg.popup_error(f"Erro ao abrir o arquivo: {e}")
        return None

# Função para salvar um arquivo Excel
def salvar_arquivo_excel(workbook, caminho_arquivo):
    try:
        if not caminho_arquivo.lower().endswith('.xlsx'):
            sg.popup_error("O arquivo deve ter a extensão .xlsx")
            return
        workbook.save(caminho_arquivo)
    except Exception as e:
        sg.popup_error(f"Erro ao salvar o arquivo: {e}")

# Função para coletar dados das tabelas
def processar_e_salvar_dados(sheet_origem, sheet_destino):
    row_index_destino = 1  # Linha inicial na planilha de destino
    tabela = None  # Nome da tabela atual

    for row in range(1, 5001):  # Ajuste o limite conforme necessário
        for column in range(1, 31):
            celula = sheet_origem.cell(row, column)
            valor_string = str(celula.value).strip().upper() if celula.value else ""

            # Identifica o título da tabela
            if valor_string in ["RECURSOS HUMANOS", "SERVIÇO DE TERCEIROS", "SERVIÇOS DE TERCEIROS",
                                "MATERIAL PERMANENTE", "MATERIAL E EQUIPAMENTO", "MATERIAL DE CONSUMO",
                                "VIAGENS E DIÁRIAS", "OUTROS"]:
                tabela = valor_string
                continue  # Passa para a próxima célula
            
            # Se uma tabela foi identificada e o valor não está vazio
            if tabela and valor_string:
                # Pular a linha de título e pegar os dados da tabela
                row_dados = row
                while celula.value and celula.value != "":
                    # Identifica as colunas de dados relevantes
                    beneficiado = str(sheet_origem.cell(row_dados, column).value) if column == 1 else ""
                    data = str(sheet_origem.cell(row_dados, column).value) if column == 2 else ""
                    cnpj = str(sheet_origem.cell(row_dados, column).value) if column == 3 else ""
                    valor = str(sheet_origem.cell(row_dados, column).value) if column == 4 else ""
                    tipo_documento = str(sheet_origem.cell(row_dados, column).value) if column == 5 else ""
                    numero_documento = str(sheet_origem.cell(row_dados, column).value) if column == 6 else ""

                    # Adiciona os dados à planilha de destino
                    if beneficiado or data or cnpj or valor or tipo_documento or numero_documento:
                        sheet_destino.append([tabela, beneficiado, data, cnpj, valor, tipo_documento, numero_documento])

                    row_dados += 1
                    celula = sheet_origem.cell(row_dados, column)

def main():
    janela_principal = [  
        [sg.Text("Qual seria o nome do arquivo origem? (coloque o caminho)")],
        [sg.InputText(key="arquivo_Origem")],
        [sg.Text("Qual seria o nome do arquivo destino? (coloque o caminho)")],
        [sg.InputText(key="arquivo_Destino")],
        [sg.Button("P&D")], [sg.Button("PE")], [sg.Button("Cancelar")] 
    ]

    window_principal = sg.Window("Dados", janela_principal)

    while True:
        event_principal, values_principal = window_principal.read()
        window_principal.close()

        if event_principal == sg.WIN_CLOSED or event_principal == 'Cancelar':
            break

        arquivoDeOrigem = values_principal["arquivo_Origem"].strip().replace('"', '')
        arquivoDeDestino = values_principal["arquivo_Destino"].strip().replace('"', '')

        if event_principal == 'PE':
            print("Chegou no PE")

            workbookOrigem = abrir_arquivo_excel(arquivoDeOrigem)
            if not workbookOrigem:
                continue

            janela_escolha = [
                [sg.Text("Qual seria a aba?")],
                [sg.Text(workbookOrigem.sheetnames)],
                [sg.InputText(key='escolha_Aba')],
                [sg.Button('Usar esta aba!')], [sg.Button('Cancelar')]
            ]

            window_escolha = sg.Window("Escolha", janela_escolha)

            while True:
                event_escolha, values_escolha = window_escolha.read()
                window_escolha.close()

                if event_escolha == sg.WIN_CLOSED or event_escolha == 'Cancelar':
                    break

                if event_escolha == 'Usar esta aba!':
                    abaEscolhida = values_escolha['escolha_Aba']
                    if abaEscolhida not in workbookOrigem.sheetnames:
                        sg.popup_error("A aba escolhida não existe.")
                        continue

                    sheet_origem = workbookOrigem[abaEscolhida]

                    # Abrir ou criar o arquivo de destino
                    workbookDestino = abrir_arquivo_excel(arquivoDeDestino)
                    if not workbookDestino:
                        workbookDestino = openpyxl.Workbook()
                    
                    sheet_destino = workbookDestino.active
                    sheet_destino.title = "Dados Processados"

                    # Adiciona cabeçalhos na planilha de destino
                    cabeçalhos = ["Tabela", "Beneficiado", "Data", "CNPJ", "Valor", "Tipo de Documento", "Número Documento"]
                    sheet_destino.append(cabeçalhos)

                    processar_e_salvar_dados(sheet_origem, sheet_destino)
                    salvar_arquivo_excel(workbookDestino, arquivoDeDestino)
                    sg.popup("Dados processados e salvos com sucesso!")
                    break

if __name__ == "__main__":
    main()
