import FreeSimpleGUI as sg
import openpyxl
from openpyxl import load_workbook

sg.theme('Black')

# Função para abrir um arquivo Excel
def abrir_arquivo_excel(caminho_arquivo):
    try:
        if not caminho_arquivo.lower().endswith('.xlsx'):
            sg.popup_error("O arquivo deve ter a extensão .xlsx")
            return None
        return load_workbook(filename=caminho_arquivo)
    except Exception as e:
        sg.popup_error(f"Erro ao abrir o arquivo: {e}")
        return None

# Função para salvar um arquivo Excel
def salvar_arquivo_excel(workbook, caminho_arquivo):
    try:
        if not caminho_arquivo.lower().endswith('.xlsx'):
            sg.popup_error("O arquivo deve ter a extensão .xlsx")
            return
        workbook.save(caminho_arquivo)
    except Exception as e:
        sg.popup_error(f"Erro ao salvar o arquivo: {e}")

# Função para ler dados da planilha de origem
def ler_dados_origem(sheet):
    dados = []
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            celula = sheet.cell(row, column)
            if celula.value:
                valor_string = str(celula.value).upper()
                if valor_string in [
                    "RECURSOS HUMANOS", "RECURSO HUMANO", "SERVIÇO DE TERCEIROS",
                    "SERVIÇOS DE TERCEIROS", "MATERIAL PERMANENTE", "MATERIAL E EQUIPAMENTO",
                    "MATERIAL DE CONSUMO", "VIAGENS E DIARIAS", "VIAGENS E DIÁRIAS", "OUTROS"
                ]:
                    dados.append((valor_string, row, column))
    return dados

# Função para inserir dados na planilha de destino
def inserir_dados_destino(sheet, dados):
    dados.sort(key=lambda x: x[0])
    row_index = 1
    for valor_string, row, column in dados:
        sheet.cell(row=row_index, column=1, value=valor_string)
        sheet.cell(row=row_index, column=2, value=row)
        sheet.cell(row=row_index, column=3, value=column)
        row_index += 1

def main():
    janela_principal = [  
        [sg.Text("Qual seria o nome do arquivo origem? (coloque o caminho)")],
        [sg.InputText(key="arquivo_Origem")],
        [sg.Text("Qual seria o nome do arquivo destino? (coloque o caminho)")],
        [sg.InputText(key="arquivo_Destino")],
        [sg.Button("P&D")], [sg.Button("PE")], [sg.Button("Cancelar")] 
    ]

    window_principal = sg.Window("Dados", janela_principal)

    while True:
        event_principal, values_principal = window_principal.read()
        window_principal.close()

        if event_principal == sg.WIN_CLOSED or event_principal == 'Cancelar':
            break

        arquivoDeOrigem = values_principal["arquivo_Origem"].strip().replace('"', '')
        arquivoDeDestino = values_principal["arquivo_Destino"].strip().replace('"', '')

        if event_principal == 'P&D':
            print("Chegou no P&D")
            
            janela_XML = [
                [sg.Text("Quem está descrito no XML?")],
                [sg.InputText(key="descrito_XML")],
                [sg.Button("Proseguir")], [sg.Button("Cancelar")]
            ]
            
            window_XML = sg.Window("Descritos no XML", janela_XML)
            
            while True:
                event_XML, values_XML = window_XML.read()
                window_XML.close()

                if event_XML == "Cancelar":
                    janela_problema = [
                        [sg.Text("Tem certeza? Isso irá fechar o programa!")],
                        [sg.Button("Sim")], [sg.Button("Não")]
                    ]

                    window_problema = sg.Window("Está correto?", janela_problema)
            
                    while True:
                        event_problema, values_problema = window_problema.read()
                        window_problema.close()

                        if event_problema == "Não":
                            break

                        if event_problema == "Sim":
                            break

                if event_XML == "Proseguir":
                    # Lógica para P&D (se necessário)
                    break

        if event_principal == 'PE':
            print("Chegou no PE")

            # Abrir o arquivo de origem
            workbookOrigem = abrir_arquivo_excel(arquivoDeOrigem)
            if not workbookOrigem:
                continue

            janela_escolha = [
                [sg.Text("Qual seria a aba?")],
                [sg.Text(workbookOrigem.sheetnames)],
                [sg.InputText(key='escolha_Aba')],
                [sg.Button('Usar esta aba!')], [sg.Button('Cancelar')]
            ]

            window_escolha = sg.Window("Escolha", janela_escolha)

            while True:
                event_escolha, values_escolha = window_escolha.read()
                window_escolha.close()

                if event_escolha == sg.WIN_CLOSED or event_escolha == 'Cancelar':
                    break

                if event_escolha == 'Usar esta aba!':
                    abaEscolhida = values_escolha['escolha_Aba']
                    sheet = workbookOrigem[abaEscolhida]

                    dados = ler_dados_origem(sheet)

                    # Criar e preencher um novo workbook para o destino
                    workbookDestino = openpyxl.Workbook()
                    sheet_destino = workbookDestino.active
                    sheet_destino.title = "Dados Ordenados"

                    inserir_dados_destino(sheet_destino, dados)

                    salvar_arquivo_excel(workbookDestino, arquivoDeDestino)
                    sg.popup("Dados processados e salvos com sucesso!")
                    break

if __name__ == "__main__":
    main()
